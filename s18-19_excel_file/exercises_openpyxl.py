import openpyxl

# load workbook

wk = openpyxl.load_workbook(
    "/Budżet domowy 2019.xlsx")

# wszytskie nazwy kart
print(wk.sheetnames)

print("Aktywne karty: " + wk.active.title)

# tworzenie obiektu karty > a raczej przejscie na konkretną kartę

sh = wk['Luty']
print(sh.title)

# pobieranie komurki  - u mnie to nie dzila bo pokazuje funkcje liczaca
print(sh["D71"])
print(sh["D71"].value)

# rozwiazanie >> wb = openpyxl.load_workbook(filename, data_only=True)
# The data_only flag helps.

c1 = sh.cell(17, 4)  # row, column nr a nie index!!!
print(c1.value)

c2 = sh.cell(row=74, column=3)  # badz poprzez prametryzowanie
print(c2.value)
print(c2.row)
print(c2.column)

# odczyt calosci z wokrbook

# find rows info

rows = sh.max_row
columns = sh.max_column

print("rows: ", rows)
print("columns: ", columns)

for row in sh["A1":"AN257"]:  # [start - lewy gorny rog : prawy dolny rog]
    for c in row:
        print(c.value)


# for i in range(1, rows + 1):
#     for j in range(1, columns + 1):
#         c = sh.cell(i, j)
#         print(c.value)
